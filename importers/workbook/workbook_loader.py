"""
Workbook loader — opens Kirgo Numbers.xlsx once and caches all sheets.

Sheet classification:
  IMPORT_SHEETS  — commerce/logistics sheets read with header in row 0
  BANK_SHEETS    — HDFC statement sheets; header at row 20, data from row 22
  EXCLUDED_SHEETS — never read (Credentials)

All DataFrames are returned with:
  - dtype=str (all values are strings)
  - NaN converted to ''
  - Column headers: lowercased + stripped
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import pandas as pd

logger = logging.getLogger(__name__)

# ── Sheet name constants ──────────────────────────────────────────────────────

IMPORT_SHEETS: list[str] = [
    "Woocom - Orders",
    "SR - 2023",
    "SR - 2024",
    "SR - 2025",
    "SR - 2026",
    "Returns - 2023",
    "Returns - 2024",
    "Returns - 2025",
    "Returns 2025 - 2026 ",   # trailing space — exact name
]

# Note: '2025 ' has a trailing space — must use exact string
BANK_SHEETS: list[str] = ["2023", "2024", "2025 ", "2026"]

EXCLUDED_SHEETS: list[str] = ["Credentials"]

ALL_EXPECTED_SHEETS: list[str] = IMPORT_SHEETS + BANK_SHEETS + EXCLUDED_SHEETS


# ── Data structures ───────────────────────────────────────────────────────────

@dataclass
class WorkbookData:
    path: Path
    # sheet_name (exact, original case) → DataFrame
    sheets: dict[str, pd.DataFrame] = field(default_factory=dict)


class WorkbookLoadError(Exception):
    """Raised when the workbook cannot be loaded or a required sheet is missing."""


# ── Public entry point ────────────────────────────────────────────────────────

def load_workbook(path: Path) -> WorkbookData:
    """
    Open the workbook once. Read all import-target sheets into DataFrames.
    The Credentials sheet is detected and logged but never read.

    Returns WorkbookData with all sheets pre-loaded.
    Raises WorkbookLoadError on missing sheet or unreadable file.
    """
    if not path.exists():
        raise WorkbookLoadError(f"Workbook not found: {path}")

    try:
        import openpyxl
        wb_meta = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        actual_sheet_names: list[str] = wb_meta.sheetnames
        wb_meta.close()
    except Exception as exc:
        raise WorkbookLoadError(f"Cannot open workbook: {exc}") from exc

    data = WorkbookData(path=path)

    # ── Check for Credentials sheet ──────────────────────────────────────────
    for excl in EXCLUDED_SHEETS:
        if excl in actual_sheet_names:
            logger.warning(
                "credentials_sheet_detected sheet=%r — contents NOT read; "
                "credentials must be rotated and moved to a secrets manager",
                excl,
            )

    # ── Load commerce/logistics sheets ───────────────────────────────────────
    for sheet in IMPORT_SHEETS:
        if sheet not in actual_sheet_names:
            raise WorkbookLoadError(
                f"Required sheet {sheet!r} not found in workbook. "
                f"Available sheets: {actual_sheet_names}"
            )
        try:
            df = pd.read_excel(
                path,
                sheet_name=sheet,
                dtype=str,
                keep_default_na=False,
                engine="openpyxl",
            )
            df = df.fillna("")
            df.columns = [str(c).lower().strip() for c in df.columns]
            data.sheets[sheet] = df
            logger.info("sheet_loaded sheet=%r rows=%d cols=%d", sheet, len(df), len(df.columns))
        except WorkbookLoadError:
            raise
        except Exception as exc:
            raise WorkbookLoadError(f"Cannot read sheet {sheet!r}: {exc}") from exc

    # ── Load bank sheets ──────────────────────────────────────────────────────
    for sheet in BANK_SHEETS:
        if sheet not in actual_sheet_names:
            raise WorkbookLoadError(
                f"Required bank sheet {sheet!r} not found. "
                f"Available sheets: {actual_sheet_names}"
            )
        try:
            df = _read_bank_sheet(path, sheet)
            data.sheets[sheet] = df
            logger.info(
                "bank_sheet_loaded sheet=%r transactions=%d", sheet, len(df)
            )
        except WorkbookLoadError:
            raise
        except Exception as exc:
            raise WorkbookLoadError(f"Cannot read bank sheet {sheet!r}: {exc}") from exc

    return data


# ── Bank sheet reader ─────────────────────────────────────────────────────────

def _read_bank_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    """
    Read an HDFC bank statement sheet.

    Layout:
      Row 0-19: account metadata (account number, branch, period, etc.)
      Row 20:   COLUMN HEADERS  ← header=20
      Row 21:   asterisk separator  ← drop as first data row
      Row 22+:  transaction data
      Footer:   statement summary rows (non-parseable Date fields)
    """
    df = pd.read_excel(
        path,
        sheet_name=sheet_name,
        header=20,          # row index 20 is the column header
        dtype=str,
        keep_default_na=False,
        engine="openpyxl",
    )
    df = df.fillna("")
    df.columns = [str(c).lower().strip() for c in df.columns]

    # Drop first data row (the asterisk separator at original row 21)
    df = df.iloc[1:].reset_index(drop=True)

    # Filter to actual transaction rows: Date must be parseable as dd/mm/yy
    mask = df["date"].apply(_is_transaction_date)
    df = df[mask].reset_index(drop=True)

    return df


def _is_transaction_date(val: str) -> bool:
    """Return True iff val is a valid transaction date in dd/mm/yy format."""
    try:
        datetime.strptime(str(val).strip(), "%d/%m/%y")
        return True
    except ValueError:
        return False


# ── Preflight report ──────────────────────────────────────────────────────────

def preflight_report(path: Path) -> list[tuple[str, str, str]]:
    """
    Run preflight checks on the workbook without writing to the DB.
    Returns a list of (status, sheet_name, detail) tuples.
    status is one of 'PASS', 'WARN', 'FAIL'.
    """
    results: list[tuple[str, str, str]] = []

    if not path.exists():
        results.append(("FAIL", str(path), "File not found"))
        return results

    try:
        import openpyxl
        wb_meta = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        actual_names = wb_meta.sheetnames
        wb_meta.close()
    except Exception as exc:
        results.append(("FAIL", str(path), f"Cannot open workbook: {exc}"))
        return results

    for excl in EXCLUDED_SHEETS:
        if excl in actual_names:
            results.append(("WARN", excl, "Credentials sheet detected — contents NOT read"))

    try:
        data = load_workbook(path)
    except WorkbookLoadError as exc:
        results.append(("FAIL", "workbook", str(exc)))
        return results

    for sheet in IMPORT_SHEETS:
        df = data.sheets.get(sheet)
        if df is None:
            results.append(("FAIL", sheet, "Sheet not loaded"))
        else:
            results.append(("PASS", sheet, f"{len(df)} rows, {len(df.columns)} cols"))

    for sheet in BANK_SHEETS:
        df = data.sheets.get(sheet)
        if df is None:
            results.append(("FAIL", sheet, "Bank sheet not loaded"))
        else:
            results.append(("PASS", sheet, f"{len(df)} transactions"))

    return results
